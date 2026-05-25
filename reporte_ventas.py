# Importo las librerias que voy a estar utilizando
import pandas as pd
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from openpyxl.styles import Font

carpeta_proyecto = Path(__file__).parent

carpeta_ventas = carpeta_proyecto / "ventas"
carpeta_output = carpeta_proyecto / "output"

# Creo la carpeta output 
carpeta_output.mkdir(exist_ok=True)

# Creo una lista vacia para guardar las tablas de los archivos Excel
tablas = []

archivos_excel = list(carpeta_ventas.glob("*.xlsx"))

if not archivos_excel:
    print("Nose encontraron archivos Excel en la carpeta ventas.")
    exit()
columnas_requeridas = {"producto", "cantidad", "precio", "ciudad"}
# Lee todos los archivos dentro de la carpeta ventas
for archivo in archivos_excel:
    df = pd.read_excel(archivo)

    columnas_archivo = set(df.columns)
    if not columnas_requeridas.issubset(columnas_archivo):
        print(f"El archivo {archivo.name} no tiene las columnas requeridas.")
        print(f"Columnas requeridas: {columnas_requeridas}")
        print(f"Columnas encontradas {columnas_archivo}")
        exit()

    df["archivo_origen"] = archivo.name

    #Guardo las tablas dentro de la lista
    tablas.append(df)

# Uno las tablas en una tabla nueva
df_final = pd.concat(tablas, ignore_index= True)

# Limpio las columnas de texto
df_final["producto"] = df_final["producto"].str.strip().str.lower()
df_final["precio"] = pd.to_numeric(df_final["precio"], errors="coerce")

# Elimino filas con datos importantes vacios o invalidos
df_final = df_final.dropna(subset = ["producto","cantidad","precio","ciudad"])

# Creo una nueva columna con el total
df_final["total"] = df_final["cantidad"] * df_final["precio"]

# Creo un resumen por ciudad y producto
resumen_ciudad = (
    df_final.groupby("ciudad")["total"]
    .sum()
    .reset_index()
    .sort_values(by="total",ascending=False))
resumen_producto = (
    df_final.groupby("producto")["total"]
    .sum()
    .reset_index()
    .sort_values(by="total",ascending=False))

# Ruta del reporte
archivo_salida = carpeta_output / "reporte_final.xlsx"

# Creo archivo Excel con varias hojas
with pd.ExcelWriter(archivo_salida, engine="openpyxl") as writer:
    df_final.to_excel(writer,sheet_name = "Datos completos", index=False)
    resumen_ciudad.to_excel(writer,sheet_name = "Resumen por ciudad", index=False)
    resumen_producto.to_excel(writer,sheet_name = "Resumen por producto", index=False)

    # Ajusto el ancho de las columnas en Excel
    for hoja in writer.sheets.values():
        for columna in hoja.columns:
            largo_maximo = 0
            letra_columna = get_column_letter(columna[0].column)

            for celda in columna:
                if celda.value is not None:
                    largo_maximo = max(largo_maximo, len(str(celda.value)))

            hoja.column_dimensions[letra_columna].width = largo_maximo + 2

    # Dar formato de monedas a las columnas llamadas "total"
    for hoja in writer.sheets.values():
        for fila in hoja.iter_rows(min_row=1, max_row=1):
            for celda in fila:
                if celda.value == "total":
                    columna_total = celda.column

                    for celda_total in hoja.iter_cols(
                        min_col = columna_total,
                        max_col = columna_total,
                        min_row = 2
                    ):
                        for celda_valor in celda_total:
                            celda_valor.number_format = '$#,##0.00'

    # Dar más ancho a la columna total para que se vea el formato de moneda
    for hoja in writer.sheets.values():
        for fila in hoja.iter_rows(min_row=1, max_row=1):
            for celda in fila:
                if celda.value == "total":
                    letra_columna = get_column_letter(celda.column)
                    hoja.column_dimensions[letra_columna].width = 15

    # Poner los encabezados en negrita
    for hoja in writer.sheets.values():
        for celda in hoja[1]:
            celda.font = Font(bold=True)

print(f"Reporte creado correctamente: {archivo_salida}")